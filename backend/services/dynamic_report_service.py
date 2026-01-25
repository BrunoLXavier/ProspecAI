"""
Dynamic Report Service
Implements RF-09: Customizable reports based on database structure with visual query builder support.
Supports schema introspection, dynamic query building, and multi-format export.
"""
from datetime import datetime, timedelta
from typing import Optional, List, Dict, Any, Tuple
from uuid import UUID, uuid4
import logging
import json
import io
import csv

from sqlalchemy import select, update, text, inspect, MetaData, Table
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.engine import Engine

from adapters.database.models import (
    ReportTemplateModel, ReportInstanceModel, ReportableTableModel,
    FundingSourceModel, ProjectModel, PortfolioModel
)

logger = logging.getLogger(__name__)


# =============================================================================
# REPORTABLE TABLES CONFIGURATION
# =============================================================================

# Define which tables are available for reporting and their metadata
REPORTABLE_TABLES: Dict[str, Dict[str, Any]] = {
    'funding_sources': {
        'display_name': 'Funding Sources',
        'description': 'Funding opportunities and grant programs',
        'fields': [
            {'name': 'id', 'type': 'uuid', 'display_name': 'ID', 'filterable': True, 'sortable': True},
            {'name': 'name', 'type': 'string', 'display_name': 'Name', 'filterable': True, 'sortable': True},
            {'name': 'description', 'type': 'text', 'display_name': 'Description', 'filterable': True, 'sortable': False},
            {'name': 'institution', 'type': 'string', 'display_name': 'Institution', 'filterable': True, 'sortable': True},
            {'name': 'instrument_type', 'type': 'string', 'display_name': 'Instrument Type', 'filterable': True, 'sortable': True},
            {'name': 'trl_min', 'type': 'integer', 'display_name': 'TRL Min', 'filterable': True, 'sortable': True},
            {'name': 'trl_max', 'type': 'integer', 'display_name': 'TRL Max', 'filterable': True, 'sortable': True},
            {'name': 'total_amount', 'type': 'decimal', 'display_name': 'Total Amount', 'filterable': True, 'sortable': True},
            {'name': 'available_amount', 'type': 'decimal', 'display_name': 'Available Amount', 'filterable': True, 'sortable': True},
            {'name': 'currency', 'type': 'string', 'display_name': 'Currency', 'filterable': True, 'sortable': True},
            {'name': 'submission_start', 'type': 'datetime', 'display_name': 'Submission Start', 'filterable': True, 'sortable': True},
            {'name': 'submission_end', 'type': 'datetime', 'display_name': 'Submission End', 'filterable': True, 'sortable': True},
            {'name': 'status', 'type': 'string', 'display_name': 'Status', 'filterable': True, 'sortable': True},
            {'name': 'source_organization', 'type': 'string', 'display_name': 'Source Organization', 'filterable': True, 'sortable': True},
            {'name': 'created_at', 'type': 'datetime', 'display_name': 'Created At', 'filterable': True, 'sortable': True},
        ],
        'relationships': [
            {'target_table': 'projects', 'join_field': 'id', 'target_field': 'funding_id', 'label': 'Related Projects', 'type': 'one_to_many'},
        ],
        'requires_permission': None,
        'display_order': 1
    },
    'projects': {
        'display_name': 'Projects',
        'description': 'Research and development projects',
        'fields': [
            {'name': 'id', 'type': 'uuid', 'display_name': 'ID', 'filterable': True, 'sortable': True},
            {'name': 'title', 'type': 'string', 'display_name': 'Title', 'filterable': True, 'sortable': True},
            {'name': 'description', 'type': 'text', 'display_name': 'Description', 'filterable': True, 'sortable': False},
            {'name': 'status', 'type': 'string', 'display_name': 'Status', 'filterable': True, 'sortable': True},
            {'name': 'trl_current', 'type': 'integer', 'display_name': 'Current TRL', 'filterable': True, 'sortable': True},
            {'name': 'trl_target', 'type': 'integer', 'display_name': 'Target TRL', 'filterable': True, 'sortable': True},
            {'name': 'research_area', 'type': 'string', 'display_name': 'Research Area', 'filterable': True, 'sortable': True},
            {'name': 'start_date', 'type': 'datetime', 'display_name': 'Start Date', 'filterable': True, 'sortable': True},
            {'name': 'end_date', 'type': 'datetime', 'display_name': 'End Date', 'filterable': True, 'sortable': True},
            {'name': 'budget', 'type': 'decimal', 'display_name': 'Budget', 'filterable': True, 'sortable': True},
            {'name': 'created_at', 'type': 'datetime', 'display_name': 'Created At', 'filterable': True, 'sortable': True},
        ],
        'relationships': [
            {'target_table': 'portfolios', 'join_field': 'portfolio_id', 'target_field': 'id', 'label': 'Portfolio', 'type': 'many_to_one'},
            {'target_table': 'institutes', 'join_field': 'institute_id', 'target_field': 'id', 'label': 'Institute', 'type': 'many_to_one'},
        ],
        'requires_permission': None,
        'display_order': 2
    },
    'portfolios': {
        'display_name': 'Portfolios',
        'description': 'Project portfolios and collections',
        'fields': [
            {'name': 'id', 'type': 'uuid', 'display_name': 'ID', 'filterable': True, 'sortable': True},
            {'name': 'name', 'type': 'string', 'display_name': 'Name', 'filterable': True, 'sortable': True},
            {'name': 'description', 'type': 'text', 'display_name': 'Description', 'filterable': True, 'sortable': False},
            {'name': 'created_at', 'type': 'datetime', 'display_name': 'Created At', 'filterable': True, 'sortable': True},
        ],
        'relationships': [
            {'target_table': 'projects', 'join_field': 'id', 'target_field': 'portfolio_id', 'label': 'Projects', 'type': 'one_to_many'},
        ],
        'requires_permission': None,
        'display_order': 3
    },
    'opportunities': {
        'display_name': 'Opportunities',
        'description': 'Business opportunities in the pipeline',
        'fields': [
            {'name': 'id', 'type': 'uuid', 'display_name': 'ID', 'filterable': True, 'sortable': True},
            {'name': 'title', 'type': 'string', 'display_name': 'Title', 'filterable': True, 'sortable': True},
            {'name': 'description', 'type': 'text', 'display_name': 'Description', 'filterable': True, 'sortable': False},
            {'name': 'stage', 'type': 'string', 'display_name': 'Stage', 'filterable': True, 'sortable': True},
            {'name': 'probability', 'type': 'decimal', 'display_name': 'Probability', 'filterable': True, 'sortable': True},
            {'name': 'expected_value', 'type': 'decimal', 'display_name': 'Expected Value', 'filterable': True, 'sortable': True},
            {'name': 'expected_close_date', 'type': 'datetime', 'display_name': 'Expected Close Date', 'filterable': True, 'sortable': True},
            {'name': 'status', 'type': 'string', 'display_name': 'Status', 'filterable': True, 'sortable': True},
            {'name': 'created_at', 'type': 'datetime', 'display_name': 'Created At', 'filterable': True, 'sortable': True},
        ],
        'relationships': [
            {'target_table': 'contacts', 'join_field': 'contact_id', 'target_field': 'id', 'label': 'Contact', 'type': 'many_to_one'},
            {'target_table': 'funding_sources', 'join_field': 'funding_source_id', 'target_field': 'id', 'label': 'Funding Source', 'type': 'many_to_one'},
        ],
        'requires_permission': 'crm:read',
        'display_order': 4
    },
    'contacts': {
        'display_name': 'Contacts',
        'description': 'CRM contacts and companies',
        'fields': [
            {'name': 'id', 'type': 'uuid', 'display_name': 'ID', 'filterable': True, 'sortable': True},
            {'name': 'name', 'type': 'string', 'display_name': 'Name', 'filterable': True, 'sortable': True},
            {'name': 'email', 'type': 'string', 'display_name': 'Email', 'filterable': True, 'sortable': True},
            {'name': 'company', 'type': 'string', 'display_name': 'Company', 'filterable': True, 'sortable': True},
            {'name': 'cnpj', 'type': 'string', 'display_name': 'CNPJ', 'filterable': True, 'sortable': True},
            {'name': 'status', 'type': 'string', 'display_name': 'Status', 'filterable': True, 'sortable': True},
            {'name': 'created_at', 'type': 'datetime', 'display_name': 'Created At', 'filterable': True, 'sortable': True},
        ],
        'relationships': [
            {'target_table': 'opportunities', 'join_field': 'id', 'target_field': 'contact_id', 'label': 'Opportunities', 'type': 'one_to_many'},
        ],
        'requires_permission': 'crm:read',
        'display_order': 5
    },
    'proposals': {
        'display_name': 'Proposals',
        'description': 'Project proposals and submissions',
        'fields': [
            {'name': 'id', 'type': 'uuid', 'display_name': 'ID', 'filterable': True, 'sortable': True},
            {'name': 'title', 'type': 'string', 'display_name': 'Title', 'filterable': True, 'sortable': True},
            {'name': 'status', 'type': 'string', 'display_name': 'Status', 'filterable': True, 'sortable': True},
            {'name': 'version', 'type': 'integer', 'display_name': 'Version', 'filterable': True, 'sortable': True},
            {'name': 'submitted_at', 'type': 'datetime', 'display_name': 'Submitted At', 'filterable': True, 'sortable': True},
            {'name': 'created_at', 'type': 'datetime', 'display_name': 'Created At', 'filterable': True, 'sortable': True},
        ],
        'relationships': [
            {'target_table': 'funding_sources', 'join_field': 'funding_source_id', 'target_field': 'id', 'label': 'Funding Source', 'type': 'many_to_one'},
            {'target_table': 'projects', 'join_field': 'project_id', 'target_field': 'id', 'label': 'Project', 'type': 'many_to_one'},
        ],
        'requires_permission': None,
        'display_order': 6
    },
    'matching_results': {
        'display_name': 'Matching Results',
        'description': 'AI-generated matching between projects and funding sources',
        'fields': [
            {'name': 'id', 'type': 'uuid', 'display_name': 'ID', 'filterable': True, 'sortable': True},
            {'name': 'overall_score', 'type': 'decimal', 'display_name': 'Overall Score', 'filterable': True, 'sortable': True},
            {'name': 'technical_score', 'type': 'decimal', 'display_name': 'Technical Score', 'filterable': True, 'sortable': True},
            {'name': 'financial_score', 'type': 'decimal', 'display_name': 'Financial Score', 'filterable': True, 'sortable': True},
            {'name': 'strategic_score', 'type': 'decimal', 'display_name': 'Strategic Score', 'filterable': True, 'sortable': True},
            {'name': 'status', 'type': 'string', 'display_name': 'Status', 'filterable': True, 'sortable': True},
            {'name': 'created_at', 'type': 'datetime', 'display_name': 'Created At', 'filterable': True, 'sortable': True},
        ],
        'relationships': [
            {'target_table': 'projects', 'join_field': 'project_id', 'target_field': 'id', 'label': 'Project', 'type': 'many_to_one'},
            {'target_table': 'funding_sources', 'join_field': 'funding_source_id', 'target_field': 'id', 'label': 'Funding Source', 'type': 'many_to_one'},
        ],
        'requires_permission': 'matching:read',
        'display_order': 7
    },
}


# SQL operators for filtering
FILTER_OPERATORS = {
    'eq': '=',
    'neq': '!=',
    'gt': '>',
    'gte': '>=',
    'lt': '<',
    'lte': '<=',
    'like': 'LIKE',
    'ilike': 'ILIKE',
    'in': 'IN',
    'not_in': 'NOT IN',
    'is_null': 'IS NULL',
    'is_not_null': 'IS NOT NULL',
    'between': 'BETWEEN',
}


class DynamicReportService:
    """
    Service for building and executing dynamic reports based on database schema.
    Supports visual query building, multiple export formats, and scheduled execution.
    """

    def __init__(self, db: AsyncSession):
        self.db = db

    # =========================================================================
    # SCHEMA INTROSPECTION
    # =========================================================================

    def get_reportable_tables(self, user_permissions: Optional[List[str]] = None) -> List[Dict[str, Any]]:
        """
        Get list of tables available for reporting.
        Filters by user permissions if provided.
        """
        tables = []
        for table_name, config in REPORTABLE_TABLES.items():
            # Check permission requirement
            required_perm = config.get('requires_permission')
            if required_perm and user_permissions:
                if required_perm not in user_permissions:
                    continue
            
            tables.append({
                'table_name': table_name,
                'display_name': config['display_name'],
                'description': config['description'],
                'field_count': len(config['fields']),
                'display_order': config['display_order']
            })
        
        return sorted(tables, key=lambda x: x['display_order'])

    def get_table_schema(self, table_name: str) -> Optional[Dict[str, Any]]:
        """Get detailed schema for a specific table."""
        config = REPORTABLE_TABLES.get(table_name)
        if not config:
            return None
        
        return {
            'table_name': table_name,
            'display_name': config['display_name'],
            'description': config['description'],
            'fields': config['fields'],
            'relationships': config['relationships'],
            'requires_permission': config.get('requires_permission')
        }

    def get_available_joins(self, base_table: str) -> List[Dict[str, Any]]:
        """Get available join options for a base table."""
        config = REPORTABLE_TABLES.get(base_table)
        if not config:
            return []
        
        joins = []
        for rel in config.get('relationships', []):
            target_config = REPORTABLE_TABLES.get(rel['target_table'])
            if target_config:
                joins.append({
                    'target_table': rel['target_table'],
                    'target_display_name': target_config['display_name'],
                    'join_field': rel['join_field'],
                    'target_field': rel['target_field'],
                    'label': rel['label'],
                    'type': rel['type'],
                    'target_fields': target_config['fields']
                })
        
        return joins

    # =========================================================================
    # REPORT TEMPLATE CRUD
    # =========================================================================

    async def create_template(
        self,
        tenant_id: UUID,
        name: str,
        query_config: Dict[str, Any],
        created_by: UUID,
        description: Optional[str] = None,
        visibility: str = 'private',
        institute_id: Optional[UUID] = None,
        display_config: Optional[Dict[str, Any]] = None,
        output_formats: Optional[List[str]] = None,
        schedule_cron: Optional[str] = None,
        schedule_enabled: bool = False,
        schedule_recipients: Optional[List[str]] = None,
        category: Optional[str] = None,
        tags: Optional[List[str]] = None
    ) -> ReportTemplateModel:
        """Create a new report template."""
        # Validate query config
        self._validate_query_config(query_config)
        
        template = ReportTemplateModel(
            id=uuid4(),
            tenant_id=tenant_id,
            name=name,
            description=description,
            visibility=visibility,
            institute_id=institute_id,
            query_config=query_config,
            display_config=display_config or {},
            output_formats=output_formats or ['html', 'csv', 'json', 'pdf', 'xlsx'],
            schedule_cron=schedule_cron,
            schedule_enabled=schedule_enabled,
            schedule_recipients=schedule_recipients or [],
            category=category,
            tags=tags or [],
            created_by=created_by,
            updated_by=created_by
        )
        
        self.db.add(template)
        await self.db.commit()
        await self.db.refresh(template)
        return template

    async def get_template(self, template_id: UUID, tenant_id: UUID) -> Optional[ReportTemplateModel]:
        """Get a template by ID."""
        result = await self.db.execute(
            select(ReportTemplateModel).where(
                ReportTemplateModel.id == template_id,
                ReportTemplateModel.tenant_id == tenant_id,
                ReportTemplateModel.deleted_at.is_(None)
            )
        )
        return result.scalar_one_or_none()

    async def list_templates(
        self,
        tenant_id: UUID,
        user_id: UUID,
        institute_id: Optional[UUID] = None,
        visibility: Optional[str] = None,
        category: Optional[str] = None,
        limit: int = 50,
        offset: int = 0
    ) -> List[ReportTemplateModel]:
        """
        List report templates visible to the user.
        Includes: user's private templates, institute templates, and system templates.
        """
        from sqlalchemy import or_
        
        # Build visibility conditions
        visibility_conditions = [
            # User's own private templates
            (ReportTemplateModel.visibility == 'private') & (ReportTemplateModel.created_by == user_id),
            # System-wide templates
            ReportTemplateModel.visibility == 'all_tenants'
        ]
        
        # Institute templates if user belongs to an institute
        if institute_id:
            visibility_conditions.append(
                (ReportTemplateModel.visibility == 'institute') & 
                (ReportTemplateModel.institute_id == institute_id)
            )
        
        query = select(ReportTemplateModel).where(
            ReportTemplateModel.tenant_id == tenant_id,
            ReportTemplateModel.deleted_at.is_(None),
            or_(*visibility_conditions)
        )
        
        if visibility:
            query = query.where(ReportTemplateModel.visibility == visibility)
        
        if category:
            query = query.where(ReportTemplateModel.category == category)
        
        query = query.order_by(ReportTemplateModel.name)
        query = query.limit(limit).offset(offset)
        
        result = await self.db.execute(query)
        return list(result.scalars().all())

    async def update_template(
        self,
        template_id: UUID,
        tenant_id: UUID,
        updated_by: UUID,
        updates: Dict[str, Any]
    ) -> Optional[ReportTemplateModel]:
        """Update a report template."""
        template = await self.get_template(template_id, tenant_id)
        if not template:
            return None
        
        # Validate query_config if provided
        if 'query_config' in updates:
            self._validate_query_config(updates['query_config'])
        
        allowed_fields = [
            'name', 'description', 'visibility', 'institute_id', 'query_config',
            'display_config', 'output_formats', 'schedule_cron', 'schedule_enabled',
            'schedule_recipients', 'category', 'tags'
        ]
        
        for field in allowed_fields:
            if field in updates:
                setattr(template, field, updates[field])
        
        template.updated_by = updated_by
        template.updated_at = datetime.utcnow()
        
        await self.db.commit()
        await self.db.refresh(template)
        return template

    async def delete_template(self, template_id: UUID, tenant_id: UUID) -> bool:
        """Soft delete a report template."""
        result = await self.db.execute(
            update(ReportTemplateModel)
            .where(
                ReportTemplateModel.id == template_id,
                ReportTemplateModel.tenant_id == tenant_id
            )
            .values(deleted_at=datetime.utcnow())
        )
        await self.db.commit()
        return result.rowcount > 0

    # =========================================================================
    # REPORT GENERATION
    # =========================================================================

    async def generate_report(
        self,
        template_id: UUID,
        tenant_id: UUID,
        user_id: UUID,
        format: str,
        parameters: Optional[Dict[str, Any]] = None
    ) -> Tuple[ReportInstanceModel, Any]:
        """
        Generate a report from a template.
        Returns the report instance and the generated content.
        """
        template = await self.get_template(template_id, tenant_id)
        if not template:
            raise ValueError(f"Template {template_id} not found")
        
        if format not in template.output_formats:
            raise ValueError(f"Format {format} not supported by this template")
        
        # Create instance record
        instance = ReportInstanceModel(
            id=uuid4(),
            tenant_id=tenant_id,
            template_id=template_id,
            format=format,
            status='processing',
            parameters=parameters or {},
            started_at=datetime.utcnow(),
            created_by=user_id,
            updated_by=user_id
        )
        
        self.db.add(instance)
        await self.db.commit()
        
        try:
            # Execute the query
            query_config = template.query_config
            if parameters:
                query_config = self._apply_parameters(query_config, parameters)
            
            rows = await self._execute_query(query_config, tenant_id)
            
            # Format the output
            content = await self._format_output(rows, format, template.display_config, template.name)
            
            # Update instance
            instance.status = 'completed'
            instance.completed_at = datetime.utcnow()
            instance.row_count = len(rows)
            instance.expires_at = datetime.utcnow() + timedelta(days=7)
            
            # Update template usage stats
            template.run_count = (template.run_count or 0) + 1
            template.last_run_at = datetime.utcnow()
            
            await self.db.commit()
            await self.db.refresh(instance)
            
            return instance, content
            
        except Exception as e:
            instance.status = 'failed'
            instance.error_message = str(e)
            instance.completed_at = datetime.utcnow()
            await self.db.commit()
            raise

    async def preview_query(
        self,
        query_config: Dict[str, Any],
        tenant_id: UUID,
        limit: int = 10
    ) -> List[Dict[str, Any]]:
        """Preview query results without creating a report instance."""
        self._validate_query_config(query_config)
        
        # Force a limit for preview
        preview_config = {**query_config, 'limit': min(limit, 100)}
        
        return await self._execute_query(preview_config, tenant_id)

    # =========================================================================
    # REPORT INSTANCES
    # =========================================================================

    async def list_instances(
        self,
        tenant_id: UUID,
        user_id: Optional[UUID] = None,
        template_id: Optional[UUID] = None,
        status: Optional[str] = None,
        limit: int = 50,
        offset: int = 0
    ) -> List[ReportInstanceModel]:
        """List report instances (generated reports)."""
        query = select(ReportInstanceModel).where(
            ReportInstanceModel.tenant_id == tenant_id,
            ReportInstanceModel.deleted_at.is_(None)
        )
        
        if user_id:
            query = query.where(ReportInstanceModel.created_by == user_id)
        
        if template_id:
            query = query.where(ReportInstanceModel.template_id == template_id)
        
        if status:
            query = query.where(ReportInstanceModel.status == status)
        
        query = query.order_by(ReportInstanceModel.created_at.desc())
        query = query.limit(limit).offset(offset)
        
        result = await self.db.execute(query)
        return list(result.scalars().all())

    async def get_instance(self, instance_id: UUID, tenant_id: UUID) -> Optional[ReportInstanceModel]:
        """Get a report instance by ID."""
        result = await self.db.execute(
            select(ReportInstanceModel).where(
                ReportInstanceModel.id == instance_id,
                ReportInstanceModel.tenant_id == tenant_id,
                ReportInstanceModel.deleted_at.is_(None)
            )
        )
        return result.scalar_one_or_none()

    async def delete_instance(self, instance_id: UUID, tenant_id: UUID) -> bool:
        """Soft delete a report instance."""
        result = await self.db.execute(
            update(ReportInstanceModel)
            .where(
                ReportInstanceModel.id == instance_id,
                ReportInstanceModel.tenant_id == tenant_id
            )
            .values(deleted_at=datetime.utcnow())
        )
        await self.db.commit()
        return result.rowcount > 0

    # =========================================================================
    # INTERNAL HELPERS
    # =========================================================================

    def _validate_query_config(self, config: Dict[str, Any]):
        """Validate query configuration."""
        if not config.get('base_table'):
            raise ValueError("Query config must specify a base_table")
        
        base_table = config['base_table']
        if base_table not in REPORTABLE_TABLES:
            raise ValueError(f"Unknown table: {base_table}")
        
        # Validate selected fields
        table_config = REPORTABLE_TABLES[base_table]
        valid_fields = {f['name'] for f in table_config['fields']}
        
        for field in config.get('selected_fields', []):
            # Handle qualified names (table.field)
            field_name = field.split('.')[-1] if '.' in field else field
            if field_name not in valid_fields and field_name != '*':
                logger.warning(f"Unknown field {field} in query config")
        
        # Validate joins
        valid_joins = {r['target_table'] for r in table_config.get('relationships', [])}
        for join in config.get('joins', []):
            if join.get('table') not in valid_joins:
                raise ValueError(f"Invalid join table: {join.get('table')}")

    def _apply_parameters(
        self,
        query_config: Dict[str, Any],
        parameters: Dict[str, Any]
    ) -> Dict[str, Any]:
        """Apply runtime parameters to query config."""
        config = json.loads(json.dumps(query_config))  # Deep copy
        
        # Replace parameter placeholders in filters
        for filter_def in config.get('filters', []):
            value = filter_def.get('value')
            if isinstance(value, str) and value.startswith('$'):
                param_name = value[1:]
                if param_name in parameters:
                    filter_def['value'] = parameters[param_name]
        
        return config

    async def _execute_query(
        self,
        config: Dict[str, Any],
        tenant_id: UUID
    ) -> List[Dict[str, Any]]:
        """Execute a dynamic query and return results."""
        base_table = config['base_table']
        selected_fields = config.get('selected_fields', ['*'])
        filters = config.get('filters', [])
        joins = config.get('joins', [])
        order_by = config.get('order_by', [])
        limit = config.get('limit', 1000)
        group_by = config.get('group_by', [])
        
        # Build SQL
        if '*' in selected_fields or not selected_fields:
            select_clause = f"{base_table}.*"
        else:
            select_clause = ', '.join(
                f"{base_table}.{f}" if '.' not in f else f
                for f in selected_fields
            )
        
        sql = f"SELECT {select_clause} FROM {base_table}"
        
        # Add joins
        for join in joins:
            join_table = join['table']
            join_on = join.get('on', {})
            join_type = join.get('type', 'LEFT').upper()
            
            if join_on:
                on_clauses = [f"{k} = {v}" for k, v in join_on.items()]
                sql += f" {join_type} JOIN {join_table} ON {' AND '.join(on_clauses)}"
        
        # Add WHERE clause (always filter by tenant_id)
        where_clauses = [f"{base_table}.tenant_id = :tenant_id"]
        where_clauses.append(f"{base_table}.deleted_at IS NULL")
        
        params = {'tenant_id': str(tenant_id)}
        
        for i, filter_def in enumerate(filters):
            field = filter_def['field']
            operator = filter_def.get('operator', 'eq')
            value = filter_def.get('value')
            
            sql_op = FILTER_OPERATORS.get(operator, '=')
            param_name = f"param_{i}"
            
            if operator in ('is_null', 'is_not_null'):
                where_clauses.append(f"{field} {sql_op}")
            elif operator == 'in':
                placeholders = ', '.join(f":param_{i}_{j}" for j in range(len(value)))
                where_clauses.append(f"{field} IN ({placeholders})")
                for j, v in enumerate(value):
                    params[f"param_{i}_{j}"] = v
            elif operator == 'between':
                where_clauses.append(f"{field} BETWEEN :param_{i}_start AND :param_{i}_end")
                params[f"param_{i}_start"] = value[0]
                params[f"param_{i}_end"] = value[1]
            else:
                if operator in ('like', 'ilike'):
                    value = f"%{value}%"
                where_clauses.append(f"{field} {sql_op} :{param_name}")
                params[param_name] = value
        
        sql += f" WHERE {' AND '.join(where_clauses)}"
        
        # Add GROUP BY
        if group_by:
            sql += f" GROUP BY {', '.join(group_by)}"
        
        # Add ORDER BY
        if order_by:
            order_clauses = []
            for ob in order_by:
                field = ob['field']
                direction = ob.get('direction', 'asc').upper()
                order_clauses.append(f"{field} {direction}")
            sql += f" ORDER BY {', '.join(order_clauses)}"
        
        # Add LIMIT
        sql += f" LIMIT {min(int(limit), 10000)}"
        
        logger.debug(f"Executing report query: {sql}")
        
        # Execute query
        result = await self.db.execute(text(sql), params)
        rows = result.mappings().all()
        
        return [dict(row) for row in rows]

    async def _format_output(
        self,
        rows: List[Dict[str, Any]],
        format: str,
        display_config: Dict[str, Any],
        report_name: str
    ) -> Any:
        """Format query results in the requested format."""
        if format == 'json':
            return {
                'report_name': report_name,
                'generated_at': datetime.utcnow().isoformat(),
                'row_count': len(rows),
                'data': self._serialize_rows(rows)
            }
        
        elif format == 'csv':
            if not rows:
                return ''
            
            output = io.StringIO()
            writer = csv.DictWriter(output, fieldnames=rows[0].keys())
            writer.writeheader()
            for row in rows:
                writer.writerow(self._serialize_row(row))
            return output.getvalue()
        
        elif format == 'html':
            return self._generate_html(rows, report_name, display_config)
        
        elif format == 'xlsx':
            return await self._generate_xlsx(rows, report_name)
        
        elif format == 'pdf':
            html_content = self._generate_html(rows, report_name, display_config)
            return await self._generate_pdf(html_content)
        
        else:
            raise ValueError(f"Unsupported format: {format}")

    def _serialize_rows(self, rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Serialize rows for JSON output."""
        return [self._serialize_row(row) for row in rows]

    def _serialize_row(self, row: Dict[str, Any]) -> Dict[str, Any]:
        """Serialize a single row, converting non-JSON-serializable types."""
        result = {}
        for key, value in row.items():
            if isinstance(value, datetime):
                result[key] = value.isoformat()
            elif isinstance(value, UUID):
                result[key] = str(value)
            elif hasattr(value, '__dict__'):
                result[key] = str(value)
            else:
                result[key] = value
        return result

    def _generate_html(
        self,
        rows: List[Dict[str, Any]],
        report_name: str,
        display_config: Dict[str, Any]
    ) -> str:
        """Generate HTML report."""
        if not rows:
            return f"""
            <!DOCTYPE html>
            <html>
            <head>
                <title>{report_name}</title>
                <style>
                    body {{ font-family: Arial, sans-serif; margin: 20px; }}
                    h1 {{ color: #333; }}
                </style>
            </head>
            <body>
                <h1>{report_name}</h1>
                <p>No data found.</p>
                <p><small>Generated at: {datetime.utcnow().isoformat()}</small></p>
            </body>
            </html>
            """
        
        headers = rows[0].keys()
        
        header_row = ''.join(f'<th>{h}</th>' for h in headers)
        
        data_rows = ''
        for row in rows:
            cells = ''.join(
                f'<td>{self._format_cell(v)}</td>'
                for v in row.values()
            )
            data_rows += f'<tr>{cells}</tr>'
        
        return f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>{report_name}</title>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 20px; background: #f5f5f5; }}
                h1 {{ color: #1e40af; margin-bottom: 20px; }}
                .meta {{ color: #666; font-size: 12px; margin-bottom: 20px; }}
                table {{ 
                    width: 100%; 
                    border-collapse: collapse; 
                    background: white;
                    box-shadow: 0 1px 3px rgba(0,0,0,0.1);
                }}
                th {{ 
                    background: #1e40af; 
                    color: white; 
                    padding: 12px 8px; 
                    text-align: left;
                    font-weight: 600;
                }}
                td {{ 
                    padding: 10px 8px; 
                    border-bottom: 1px solid #e5e7eb;
                }}
                tr:hover {{ background: #f9fafb; }}
                tr:nth-child(even) {{ background: #f3f4f6; }}
                tr:nth-child(even):hover {{ background: #e5e7eb; }}
                .footer {{
                    margin-top: 20px;
                    padding: 10px;
                    color: #666;
                    font-size: 11px;
                }}
            </style>
        </head>
        <body>
            <h1>{report_name}</h1>
            <div class="meta">
                Total records: {len(rows)} | Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')}
            </div>
            <table>
                <thead><tr>{header_row}</tr></thead>
                <tbody>{data_rows}</tbody>
            </table>
            <div class="footer">
                ProspecAI Report System - RF-09
            </div>
        </body>
        </html>
        """

    def _format_cell(self, value: Any) -> str:
        """Format a cell value for HTML display."""
        if value is None:
            return '—'
        if isinstance(value, datetime):
            return value.strftime('%Y-%m-%d %H:%M')
        if isinstance(value, bool):
            return '✓' if value else '✗'
        if isinstance(value, UUID):
            return str(value)[:8] + '...'
        return str(value)

    async def _generate_xlsx(self, rows: List[Dict[str, Any]], report_name: str) -> bytes:
        """Generate XLSX report using openpyxl."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise ValueError("XLSX export requires openpyxl. Install with: pip install openpyxl")
        
        wb = Workbook()
        ws = wb.active
        ws.title = report_name[:31]  # Excel sheet name limit
        
        if not rows:
            ws['A1'] = 'No data'
        else:
            # Header row
            headers = list(rows[0].keys())
            header_fill = PatternFill(start_color='1E40AF', end_color='1E40AF', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True)
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            
            # Data rows
            for row_idx, row in enumerate(rows, 2):
                for col_idx, (key, value) in enumerate(row.items(), 1):
                    if isinstance(value, datetime):
                        value = value.replace(tzinfo=None)  # Remove timezone for Excel
                    elif isinstance(value, UUID):
                        value = str(value)
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Auto-fit columns
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[column].width = min(max_length + 2, 50)
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    async def _generate_pdf(self, html_content: str) -> bytes:
        """Generate PDF from HTML using weasyprint."""
        try:
            from weasyprint import HTML
        except ImportError:
            raise ValueError("PDF export requires weasyprint. Install with: pip install weasyprint")
        
        pdf = HTML(string=html_content).write_pdf()
        return pdf
